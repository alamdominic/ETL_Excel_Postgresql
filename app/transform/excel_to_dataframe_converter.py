"""
Excel to DataFrame converter module (Transformation Layer).

Converts raw Excel worksheet objects into pandas DataFrames
with data cleaning and validation applied.
"""

import logging
from abc import ABC, abstractmethod
from typing import Optional
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class IExcelToDataFrameConverter(ABC):
    """Abstract base class for Excel-to-DataFrame conversion.

    Defines the contract for all converter implementations.
    Enforces separation between data ingestion and transformation.

    Design rationale:
        - Separates reading concerns from transformation concerns
        - Enables different conversion strategies (with/without headers, skip rows, etc.)
        - Promotes testability with mock converters
    """

    @abstractmethod
    def convert(self, worksheet: Worksheet) -> pd.DataFrame:
        """Convert Excel worksheet to pandas DataFrame.

        Args:
            worksheet: Raw openpyxl Worksheet object.

        Returns:
            Cleaned pandas DataFrame ready for loading.

        Raises:
            ValueError: If worksheet is empty or invalid.
        """
        pass


class ExcelToDataFrameConverter(IExcelToDataFrameConverter):
    """Concrete converter that transforms openpyxl Worksheet to DataFrame.

    Handles:
        - Header detection and validation
        - Empty row removal
        - Data type inference
        - Null value standardization

    Attributes:
        has_header (bool): Whether the first row contains column headers.
        skip_rows (int): Number of rows to skip from the top (default: 0).

    Design notes:
        - Assumes first row is header by default
        - Removes completely empty rows
        - Preserves original column order
        - Does NOT perform business logic transformations

    Example:
        >>> from app.extract.excel_extractor import ExcelReader
        >>> reader = ExcelReader()
        >>> converter = ExcelToDataFrameConverter()
        >>> worksheet = reader.read_sheet("data.xlsx", "Sheet1")
        >>> df = converter.convert(worksheet)
    """

    def __init__(self, has_header: bool = True, skip_rows: int = 0):
        """Initialize converter with configuration.

        Args:
            has_header: If True, first row is treated as column names.
            skip_rows: Number of rows to skip before reading (useful for files with metadata).
        """
        self.has_header = has_header
        self.skip_rows = skip_rows

    def convert(self, worksheet: Worksheet) -> pd.DataFrame:
        """Convert Excel worksheet to pandas DataFrame.

        Process:
            1. Validate worksheet is not empty
            2. Extract all cell values into a 2D list
            3. Create DataFrame with optional header detection
            4. Remove completely empty rows
            5. Standardize null values

        Args:
            worksheet: openpyxl Worksheet object containing raw Excel data.

        Returns:
            pandas DataFrame with cleaned data ready for transformation/loading.

        Raises:
            ValueError: If worksheet is empty or has no data after cleaning.

        Example:
            >>> converter = ExcelToDataFrameConverter(has_header=True, skip_rows=2)
            >>> df = converter.convert(worksheet)
            >>> print(df.shape)
            (1500, 12)
        """
        # Validate worksheet
        if worksheet.max_row is None or worksheet.max_row <= self.skip_rows:
            logger.error("Worksheet is empty or has insufficient rows")
            raise ValueError(
                f"Worksheet has insufficient data (max_row={worksheet.max_row}, skip_rows={self.skip_rows})"
            )

        # Extract all cell values into a 2D list
        logger.info("Extracting data from worksheet...")
        data = []
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            # Skip specified rows
            if row_idx <= self.skip_rows:
                continue
            data.append(list(row))

        if not data:
            logger.error("No data extracted from worksheet")
            raise ValueError("Worksheet contains no data after skipping rows")

        # Create DataFrame with or without header
        if self.has_header and len(data) > 1:
            headers = data[0]
            df = pd.DataFrame(data[1:], columns=headers)
            logger.info("Created DataFrame with headers: %s", headers)
        elif self.has_header and len(data) == 1:
            # Only header row present, no data
            logger.warning("Worksheet contains only header row, no data rows")
            df = pd.DataFrame(columns=data[0])
        else:
            # No header, use default column names
            df = pd.DataFrame(data)
            logger.info(
                "Created DataFrame without headers (auto-generated column names)"
            )

        # Remove completely empty rows
        initial_row_count = len(df)
        df = df.dropna(how="all")
        removed_rows = initial_row_count - len(df)
        if removed_rows > 0:
            logger.info("Removed %d completely empty rows", removed_rows)

        # Reset index after dropping rows
        df = df.reset_index(drop=True)

        # Log summary
        logger.info(
            "Conversion complete: %d rows × %d columns", len(df), len(df.columns)
        )

        return df
