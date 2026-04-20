"""
Excel file reader module (Ingestion Layer).

Provides abstraction for reading Excel files from disk.
This class handles file I/O and basic validation without
performing any data transformation.
"""

import logging
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class IExcelReader(ABC):
    """Abstract base class for Excel reading operations.

    Defines the contract for all Excel reader implementations.
    This enforces consistency across different Excel reading strategies
    and promotes dependency inversion.

    Design rationale:
        - Decouples business logic from specific Excel library implementations
        - Enables testing with mock readers
        - Allows swapping libraries (openpyxl, xlrd, etc.) without changing consumer code
    """

    @abstractmethod
    def read_sheet(self, file_path: str, sheet_name: str) -> Worksheet:
        """Read a specific sheet from an Excel file.

        Args:
            file_path: Absolute or relative path to the Excel file.
            sheet_name: Name of the sheet to read.

        Returns:
            Raw worksheet object from the underlying library.

        Raises:
            FileNotFoundError: If the Excel file does not exist.
            ValueError: If the sheet name is not found in the workbook.
        """
        pass


class ExcelReader(IExcelReader):
    """Concrete Excel reader using openpyxl library.

    Handles reading .xlsx files and basic validation.
    Does NOT perform data transformation or cleaning.

    Attributes:
        None (stateless reader)

    Design notes:
        - Uses openpyxl for modern .xlsx format support
        - Validates file existence before attempting to read
        - Validates sheet name existence in workbook
        - Returns raw worksheet object for downstream processing

    Example:
        >>> reader = ExcelReader()
        >>> worksheet = reader.read_sheet("data/sales.xlsx", "Sheet1")
        >>> print(worksheet.max_row)
    """

    def read_sheet(self, file_path: str, sheet_name: str) -> Worksheet:
        """Read a specific sheet from an Excel file.

        Args:
            file_path: Path to the Excel file (relative or absolute).
            sheet_name: Name of the sheet to extract.

        Returns:
            openpyxl Worksheet object containing raw cell data.

        Raises:
            FileNotFoundError: If file_path does not exist.
            ValueError: If sheet_name is not found in the workbook.

        Example:
            >>> reader = ExcelReader()
            >>> ws = reader.read_sheet("sales.xlsx", "Q1_Data")
        """
        # Validate file existence
        file = Path(file_path)
        if not file.exists():
            logger.error("Excel file not found: %s", file_path)
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        # Load workbook
        try:
            workbook = openpyxl.load_workbook(filename=file_path, data_only=True)
            logger.info("Loaded workbook: %s", file_path)
        except Exception as e:
            logger.error("Failed to load workbook %s: %s", file_path, e)
            raise ValueError(f"Failed to load Excel file: {e}")

        # Validate sheet existence
        if sheet_name not in workbook.sheetnames:
            available_sheets = ", ".join(workbook.sheetnames)
            logger.error(
                "Sheet '%s' not found in workbook. Available sheets: %s",
                sheet_name,
                available_sheets,
            )
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {available_sheets}"
            )

        # Extract and return worksheet
        worksheet = workbook[sheet_name]
        logger.info(
            "Extracted sheet '%s' with %d rows and %d columns",
            sheet_name,
            worksheet.max_row,
            worksheet.max_column,
        )

        return worksheet
